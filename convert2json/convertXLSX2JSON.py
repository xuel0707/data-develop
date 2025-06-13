import os
import json
from openpyxl import load_workbook
from tqdm import tqdm
import argparse


def extract_sheet_from_xlsx(xlsx_path):
    """
    提取指定 .xlsx 文件中每个非空的 sheet。
    每个 sheet 返回一个字典对象，包含：
        - text: 表格内容拼接成的文本（以 \t 分隔列）
        - filename: 来源 Excel 文件名
        - sheet_name: sheet 名称
    """
    wb = None
    try:
        wb = load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        print(f"Error loading file {xlsx_path}: {e}")
        return []

    sheets_text = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet_lines = []
        has_content = False  # 是否是非空 sheet

        for row in sheet.iter_rows(values_only=True):
            row_values = [str(cell) if cell is not None else '' for cell in row]
            line = '\t'.join(row_values)
            sheet_lines.append(line)

            # 如果该行不是全空，则认为有内容
            if any(row_values):
                has_content = True

        # 只保留有内容的 sheet
        if has_content:
            sheets_text.append({
                "text": "\n".join(sheet_lines).strip(),
                "filename": os.path.basename(xlsx_path),
                "sheet_name": sheet_name
            })

    return sheets_text


def process_directory(directory, recursive=False):
    """
    遍历目录下的所有 .xlsx 文件（支持递归），提取每个非空 sheet。
    返回值：包含所有 sheet 数据的列表。
    """
    all_sheets_data = []
    xlsx_files = []

    # 查找所有 .xlsx 文件
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.lower().endswith('.xlsx'):
                xlsx_files.append(os.path.join(root, file))
        if not recursive:
            break  # 不递归就跳出循环

    # 处理每个文件
    for file_path in tqdm(xlsx_files, desc="Processing XLSX files"):
        try:
            sheets = extract_sheet_from_xlsx(file_path)
            all_sheets_data.extend(sheets)
        except Exception as e:
            print(f"Error processing {file_path}: {e}")

    return all_sheets_data


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Convert each non-empty sheet of .xlsx files to JSON objects.')
    parser.add_argument('directory', type=str, help='Path to the directory containing .xlsx files')
    parser.add_argument('output_json', type=str, help='Path to save the output JSON file')
    parser.add_argument('--recursive', action='store_true', help='Search subdirectories recursively')

    args = parser.parse_args()

    sheets_data = process_directory(args.directory, recursive=args.recursive)

    with open(args.output_json, 'w', encoding='utf-8') as f:
        json.dump(sheets_data, f, ensure_ascii=False, indent=4)

    print(f"✅ Total non-empty sheets extracted: {len(sheets_data)}")
    print(f"📄 Output saved to: {args.output_json}")